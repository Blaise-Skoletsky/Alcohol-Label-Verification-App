import csv
import io
from pathlib import Path

# Canonical columns the importer understands, in template order.
TEMPLATE_COLUMNS = [
    "image",
    "brand_name",
    "class_type",
    "beverage_class",
    "alcohol_content",
    "net_contents",
    "name_address",
    "country_of_origin",
]

# Map tolerant header spellings onto canonical columns.
_HEADER_ALIASES = {
    "image": "image",
    "image_file": "image",
    "filename": "image",
    "file": "image",
    "label_image": "image",
    "brand": "brand_name",
    "brand_name": "brand_name",
    "class_type": "class_type",
    "class/type": "class_type",
    "class_type_designation": "class_type",
    "type": "class_type",
    "beverage_class": "beverage_class",
    "bev_class": "beverage_class",
    "class": "beverage_class",
    "alcohol_content": "alcohol_content",
    "abv": "alcohol_content",
    "alcohol": "alcohol_content",
    "net_contents": "net_contents",
    "net": "net_contents",
    "net_content": "net_contents",
    "name_address": "name_address",
    "name_and_address": "name_address",
    "name_addr": "name_address",
    "producer": "name_address",
    "country_of_origin": "country_of_origin",
    "country": "country_of_origin",
    "origin": "country_of_origin",
}


class SheetParseError(ValueError):
    pass


def _normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def _canonical(header: str) -> str | None:
    return _HEADER_ALIASES.get(_normalize_header(header))


class SheetService:
    """Parses an uploaded .csv/.xlsx of application data into normalized rows.

    Stateless: the uploaded sheet is never persisted. Only the recognized
    columns are returned; unknown columns are ignored.
    """

    def parse(self, filename: str, content: bytes) -> dict:
        extension = Path(filename or "").suffix.lower()
        if extension == ".csv":
            raw_headers, raw_rows = self._read_csv(content)
        elif extension in {".xlsx", ".xlsm"}:
            raw_headers, raw_rows = self._read_xlsx(content)
        else:
            raise SheetParseError("Please upload a .csv or .xlsx spreadsheet.")

        if not raw_headers:
            raise SheetParseError("The spreadsheet has no header row.")

        # Build header index -> canonical column.
        index_map: dict[int, str] = {}
        detected: list[str] = []
        for index, header in enumerate(raw_headers):
            canonical = _canonical(header)
            if canonical and canonical not in detected:
                index_map[index] = canonical
                detected.append(canonical)

        rows: list[dict[str, str]] = []
        for raw_row in raw_rows:
            row = {column: "" for column in TEMPLATE_COLUMNS}
            has_value = False
            for index, canonical in index_map.items():
                if index < len(raw_row):
                    value = (raw_row[index] or "").strip()
                    row[canonical] = value
                    if value:
                        has_value = True
            if has_value:
                rows.append(row)

        return {
            "columns": detected,
            "rows": rows,
            "row_count": len(rows),
        }

    def template_csv(self) -> str:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(TEMPLATE_COLUMNS)
        return buffer.getvalue()

    def _read_csv(self, content: bytes) -> tuple[list[str], list[list[str]]]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise SheetParseError("The CSV file is not valid UTF-8 text.") from exc
        reader = csv.reader(io.StringIO(text))
        records = [row for row in reader]
        if not records:
            return [], []
        return records[0], records[1:]

    def _read_xlsx(self, content: bytes) -> tuple[list[str], list[list[str]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise SheetParseError(
                "Reading .xlsx files requires the openpyxl package on the backend."
            ) from exc

        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise SheetParseError("We could not read this .xlsx file.") from exc

        try:
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            records = [
                ["" if cell is None else str(cell) for cell in row]
                for row in rows_iter
            ]
        finally:
            workbook.close()

        if not records:
            return [], []
        return records[0], records[1:]
